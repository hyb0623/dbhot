# -*- coding: utf-8 -*-
# @File : excel初识.py
# @Author : hyb
# @Time : 2022/7/25 19:30
# @Software : PyCharm
from openpyxl.styles import Font, colors, Alignment

from openpyxl import Workbook, load_workbook

# wb = Workbook() # 创建表
# sheet =wb.active
# sheet["B9"] = "178"
# sheet["B8"] = "1sa"
# sheet.append(["hyb","178","70"])
#
# sheet.title = "表一"
# wb.save("练习.xlsx")
from openpyxl.drawing import colors

wb = load_workbook(r"C:\Users\hyb\Desktop\1.xlsx") #打开表
sheet = wb.get_sheet_by_name("Sheet1")
# print(sheet["B5"])
# print(sheet["B5"].value)
#
# for cell in sheet["B5:C10"]:
#     print(cell[1].value)

# for row in sheet: # 按行循环所有数据
#     # print(row)
#     for cell in row:
#         print(cell.value,end="-")
#     print()

# for row in sheet.iter_rows(min_row=4,max_row=21,max_col=5): #按行循环
#     for cell in row:
#         print(cell.value,end="-")
#     print()

# for col in sheet.columns: # 按列循环，相当于把excel转置了
#     for cell in col:
#         print(cell.value,end="-")
#     print()

bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED,
bold=True) # 声明样式
sheet['A1'].font = bold_itatic_24_font # 给单元格设置样式

